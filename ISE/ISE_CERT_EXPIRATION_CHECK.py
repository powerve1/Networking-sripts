#!/usr/bin/env python3

import csv
import datetime
import getpass
import json
import sys
from pathlib import Path

import requests
from requests.auth import HTTPBasicAuth
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HOSTS_FILE = "ise_hosts.txt"
LOG_FILE = "ise_cert_expiration.log"
CSV_FILE = "ise_cert_expiration_report.csv"
XLSX_FILE = "ise_cert_expiration_report.xlsx"
SUMMARY_FILE = "ise_cert_expiration_summary.txt"

PORT = 443
TIMEOUT_SEC = 60

HEADERS = {
    "Accept": "application/json",
    "Content-Type": "application/json",
}


def now_ts():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def log_event(message):
    entry = f"[{now_ts()}] {message}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(entry + "\n")
    print(entry, flush=True)


def load_hosts(file_path):
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"Hosts file not found: {file_path}")

    hosts = []
    for raw in p.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "#" in line:
            line = line.split("#", 1)[0].strip()
        if line:
            hosts.append(line)

    return list(dict.fromkeys(hosts))


def safe_json(resp):
    try:
        return resp.json()
    except Exception:
        return {}


def format_error_detail(resp):
    body = safe_json(resp)

    if body:
        return f"HTTP {resp.status_code} - {json.dumps(body, ensure_ascii=False)[:1500]}"

    txt = (resp.text or "").strip()
    if txt:
        return f"HTTP {resp.status_code} - {txt[:1500]}"

    return f"HTTP {resp.status_code} - No response body"


def parse_ise_expiration_date(raw):
    if not raw:
        return None

    parts = raw.split()

    if len(parts) >= 6:
        no_tz = " ".join(parts[0:4] + parts[5:6])
        try:
            return datetime.datetime.strptime(no_tz, "%a %b %d %H:%M:%S %Y").date()
        except Exception:
            pass

    for fmt in (
        "%a %b %d %H:%M:%S %Y",
        "%b %d %H:%M:%S %Y",
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
    ):
        try:
            return datetime.datetime.strptime(raw, fmt).date()
        except Exception:
            continue

    return None


def fmt_dd_mon_yyyy(d):
    if not d:
        return "UNKNOWN"
    return d.strftime("%d %b %Y")


def days_until(d):
    if not d:
        return None
    return (d - datetime.date.today()).days


def cert_status(d):
    days = days_until(d)

    if days is None:
        return "UNKNOWN"
    if days < 0:
        return "EXPIRED"
    if days <= 30:
        return "EXPIRING <=30 DAYS"
    if days <= 60:
        return "EXPIRING <=60 DAYS"
    if days <= 90:
        return "EXPIRING <=90 DAYS"

    return "OK"


def get_deployment_nodes(session, base):
    url = f"{base}/api/v1/deployment/node"

    try:
        r = session.get(url, timeout=TIMEOUT_SEC)
    except requests.exceptions.ReadTimeout:
        return None, f"Read timed out after {TIMEOUT_SEC} seconds"
    except requests.exceptions.ConnectTimeout:
        return None, f"Connection timed out after {TIMEOUT_SEC} seconds"
    except requests.exceptions.ConnectionError:
        return None, "Connection error. Check network, ACL, DNS, or port 443"

    if r.status_code == 200:
        body = safe_json(r)
        nodes = body.get("response")

        if isinstance(nodes, list):
            return nodes, "OK"

        return None, (
            "Unexpected payload from /api/v1/deployment/node. "
            "This usually means API Gateway/OpenAPI is disabled or the node returned a different page. "
            f"Raw response: {json.dumps(body, ensure_ascii=False)[:1500]}"
        )

    if r.status_code == 401:
        return None, "Unauthorized. Check OpenAPI username/password and admin role."

    if r.status_code == 403:
        return None, "Forbidden. User may not have OpenAPI permissions."

    if r.status_code == 404:
        return None, "API endpoint not found. API Gateway/OpenAPI may be disabled."

    return None, format_error_detail(r)


def resolve_node_hostname_from_ip(nodes, host_ip_or_name):
    needle = host_ip_or_name.strip().lower()

    for n in nodes:
        ip = str(n.get("ipAddress", "")).strip().lower()
        hn = str(n.get("hostname", "")).strip()

        if ip and ip == needle and hn:
            return hn

    return None


def get_system_certs_for_node(session, base, node_hostname):
    url = f"{base}/api/v1/certs/system-certificate/{node_hostname}"

    try:
        r = session.get(url, timeout=TIMEOUT_SEC)
    except requests.exceptions.ReadTimeout:
        return None, f"Read timed out after {TIMEOUT_SEC} seconds"
    except requests.exceptions.ConnectTimeout:
        return None, f"Connection timed out after {TIMEOUT_SEC} seconds"
    except requests.exceptions.ConnectionError:
        return None, "Connection error. Check network, ACL, DNS, or port 443"

    if r.status_code == 200:
        body = safe_json(r)
        certs = body.get("response")

        if isinstance(certs, list):
            return certs, "OK"

        return None, (
            "Unexpected payload from system certificate API. "
            f"Raw response: {json.dumps(body, ensure_ascii=False)[:1500]}"
        )

    if r.status_code == 401:
        return None, "Unauthorized. Check OpenAPI credentials."

    if r.status_code == 403:
        return None, "Forbidden. User may not have OpenAPI permissions."

    if r.status_code == 404:
        return None, f"Not found. Node hostname '{node_hostname}' may be incorrect."

    return None, format_error_detail(r)


def find_admin_and_eap_certs(certs):
    admin_cert = None
    eap_cert = None

    for cert in certs:
        used_by = str(cert.get("usedBy") or "").lower()

        if admin_cert is None and "admin" in used_by:
            admin_cert = cert

        if eap_cert is None and ("eap authentication" in used_by or "eap" in used_by):
            eap_cert = cert

    return admin_cert, eap_cert


def sort_key(row):
    dates = []

    if row.get("admin_expiration_date"):
        dates.append(row["admin_expiration_date"])

    if row.get("eap_expiration_date"):
        dates.append(row["eap_expiration_date"])

    if dates:
        return min(dates)

    return datetime.date.max


def write_csv(rows):
    fieldnames = [
        "host",
        "node_hostname",
        "admin_cert_friendly_name",
        "admin_cert_expires",
        "admin_days_until_expiration",
        "admin_status",
        "eap_cert_friendly_name",
        "eap_cert_expires",
        "eap_days_until_expiration",
        "eap_status",
        "earliest_expiration",
    ]

    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for row in rows:
            writer.writerow({
                "host": row.get("host", ""),
                "node_hostname": row.get("node_hostname", ""),
                "admin_cert_friendly_name": row.get("admin_cert_friendly_name", ""),
                "admin_cert_expires": row.get("admin_cert_expires", ""),
                "admin_days_until_expiration": row.get("admin_days_until_expiration", ""),
                "admin_status": row.get("admin_status", ""),
                "eap_cert_friendly_name": row.get("eap_cert_friendly_name", ""),
                "eap_cert_expires": row.get("eap_cert_expires", ""),
                "eap_days_until_expiration": row.get("eap_days_until_expiration", ""),
                "eap_status": row.get("eap_status", ""),
                "earliest_expiration": row.get("earliest_expiration", ""),
            })


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def apply_status_colors(ws):
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    green = PatternFill("solid", fgColor="C6EFCE")
    gray = PatternFill("solid", fgColor="D9E1F2")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "EXPIRED":
                cell.fill = red
            elif cell.value in ("EXPIRING <=30 DAYS", "EXPIRING <=60 DAYS", "EXPIRING <=90 DAYS"):
                cell.fill = yellow
            elif cell.value == "OK":
                cell.fill = green
            elif cell.value == "UNKNOWN":
                cell.fill = gray


def write_excel(rows, failures):
    wb = Workbook()

    ws = wb.active
    ws.title = "Certificate Report"

    headers = [
        "Host",
        "Node Hostname",
        "Admin Cert Friendly Name",
        "Admin Expires",
        "Admin Days Left",
        "Admin Status",
        "EAP Cert Friendly Name",
        "EAP Expires",
        "EAP Days Left",
        "EAP Status",
        "Earliest Expiration",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("host", ""),
            row.get("node_hostname", ""),
            row.get("admin_cert_friendly_name", ""),
            row.get("admin_cert_expires", ""),
            row.get("admin_days_until_expiration", ""),
            row.get("admin_status", ""),
            row.get("eap_cert_friendly_name", ""),
            row.get("eap_cert_expires", ""),
            row.get("eap_days_until_expiration", ""),
            row.get("eap_status", ""),
            row.get("earliest_expiration", ""),
        ])

    style_header(ws)
    apply_status_colors(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("Summary")
    summary.append(["Item", "Value"])
    summary.append(["Generated", now_ts()])
    summary.append(["Total Successful Hosts", len(rows)])
    summary.append(["Total Failed Hosts", len(failures)])
    summary.append(["CSV File", CSV_FILE])
    summary.append(["Excel File", XLSX_FILE])
    summary.append(["Log File", LOG_FILE])
    style_header(summary)
    auto_width(summary)

    failed_ws = wb.create_sheet("Failures")
    failed_ws.append(["Host", "Reason"])

    for host, reason in failures:
        failed_ws.append([host, reason])

    style_header(failed_ws)
    auto_width(failed_ws)

    wb.save(XLSX_FILE)


def write_summary(rows, failures):
    lines = []
    lines.append(f"ISE Certificate Expiration Summary - {now_ts()}")
    lines.append("=" * 72)
    lines.append("")
    lines.append(f"Hosts File   : {HOSTS_FILE}")
    lines.append(f"CSV Output   : {CSV_FILE}")
    lines.append(f"Excel Output : {XLSX_FILE}")
    lines.append(f"Log File     : {LOG_FILE}")
    lines.append("")

    lines.append(f"SUCCESS ({len(rows)}):")
    if rows:
        for r in rows:
            lines.append(
                f"  - {r['host']} ({r['node_hostname']}): "
                f"Admin={r['admin_cert_expires']} [{r['admin_status']}] | "
                f"EAP={r['eap_cert_expires']} [{r['eap_status']}]"
            )
    else:
        lines.append("  (none)")

    lines.append("")
    lines.append(f"FAILED ({len(failures)}):")

    if failures:
        for host, reason in failures:
            lines.append(f"  - {host}: {reason}")
    else:
        lines.append("  (none)")

    Path(SUMMARY_FILE).write_text("\n".join(lines), encoding="utf-8")


def write_outputs(rows, failures):
    rows_sorted = sorted(rows, key=sort_key)

    write_csv(rows_sorted)
    write_excel(rows_sorted, failures)
    write_summary(rows_sorted, failures)


def main():
    print("ISE certificate expiration report script started...", flush=True)

    api_user = input("Enter ISE OpenAPI username: ").strip()
    api_pass = getpass.getpass("Enter ISE OpenAPI password: ")

    if not api_user:
        print("ERROR: OpenAPI username is required.")
        return 2

    if not api_pass:
        print("ERROR: OpenAPI password is required.")
        return 2

    try:
        ise_hosts = load_hosts(HOSTS_FILE)
    except Exception as e:
        print(f"ERROR: {e}")
        return 2

    if not ise_hosts:
        print(f"ERROR: No hosts found in {HOSTS_FILE}")
        return 2

    log_event("------------------------------------------------------------")
    log_event("Starting certificate expiration report - Admin + EAP")
    log_event(f"Loaded {len(ise_hosts)} host(s) from {HOSTS_FILE}")
    log_event("------------------------------------------------------------")

    session = requests.Session()
    session.headers.update(HEADERS)
    session.auth = HTTPBasicAuth(api_user, api_pass)
    session.verify = False

    rows = []
    failures = []

    for host in ise_hosts:
        base = f"https://{host}:{PORT}"

        try:
            nodes, nodes_reason = get_deployment_nodes(session, base)

            if not nodes:
                reason = f"Could not retrieve deployment nodes: {nodes_reason}"
                log_event(f"FAILED: {host} - {reason}")
                failures.append((host, reason))
                continue

            node_hostname = resolve_node_hostname_from_ip(nodes, host)

            if not node_hostname:
                node_hostname = host

            certs, certs_reason = get_system_certs_for_node(session, base, node_hostname)

            if not certs:
                reason = f"Could not retrieve system certs for node '{node_hostname}': {certs_reason}"
                log_event(f"FAILED: {host} - {reason}")
                failures.append((host, reason))
                continue

            admin_cert, eap_cert = find_admin_and_eap_certs(certs)

            admin_exp = parse_ise_expiration_date(
                admin_cert.get("expirationDate") if admin_cert else ""
            )

            eap_exp = parse_ise_expiration_date(
                eap_cert.get("expirationDate") if eap_cert else ""
            )

            valid_dates = [d for d in [admin_exp, eap_exp] if d]
            earliest = min(valid_dates) if valid_dates else None

            row = {
                "host": host,
                "node_hostname": node_hostname,

                "admin_cert_friendly_name": admin_cert.get("friendlyName") if admin_cert else "NOT FOUND",
                "admin_expiration_date": admin_exp,
                "admin_cert_expires": fmt_dd_mon_yyyy(admin_exp),
                "admin_days_until_expiration": days_until(admin_exp) if admin_exp else "UNKNOWN",
                "admin_status": cert_status(admin_exp),

                "eap_cert_friendly_name": eap_cert.get("friendlyName") if eap_cert else "NOT FOUND",
                "eap_expiration_date": eap_exp,
                "eap_cert_expires": fmt_dd_mon_yyyy(eap_exp),
                "eap_days_until_expiration": days_until(eap_exp) if eap_exp else "UNKNOWN",
                "eap_status": cert_status(eap_exp),

                "earliest_expiration": fmt_dd_mon_yyyy(earliest),
            }

            rows.append(row)

            log_event(
                f"SUCCESS: {host} ({node_hostname}) "
                f"- Admin expires {row['admin_cert_expires']} "
                f"- EAP expires {row['eap_cert_expires']}"
            )

        except Exception as e:
            reason = f"Unexpected error: {str(e)}"
            log_event(f"FAILED: {host} - {reason}")
            failures.append((host, reason))

    write_outputs(rows, failures)

    log_event("------------------------------------------------------------")
    log_event(f"Finished. Success: {len(rows)} | Failed: {len(failures)}")
    log_event(f"CSV report     : {CSV_FILE}")
    log_event(f"Excel report   : {XLSX_FILE}")
    log_event(f"Summary report : {SUMMARY_FILE}")
    log_event(f"Detailed log   : {LOG_FILE}")
    log_event("------------------------------------------------------------")

    return 0 if not failures else 1


if __name__ == "__main__":
    sys.exit(main())